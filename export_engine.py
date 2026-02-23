# =============================================================================
# export_engine.py — Export CSV/XLSX/PDF/SVG/HTML v7.0
# =============================================================================
# Gera ficheiros de export a partir de dados de tool results.
# CSV/XLSX: stdlib + openpyxl. PDF: fpdf2. SVG: geração manual.
# =============================================================================

import csv
import io
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

from config import EXPORT_BRAND_COLOR, EXPORT_BRAND_NAME, EXPORT_AGENT_NAME

# =============================================================================
# DATA EXTRACTION
# =============================================================================

def extract_table_data(tool_result: dict) -> tuple[List[str], List[List[str]]]:
    """Extrai headers e rows de um tool result."""
    items = tool_result.get("items", tool_result.get("analysis_data", []))
    if not items:
        # KPI groups?
        groups = tool_result.get("groups", [])
        if groups:
            return ["Valor", "Contagem"], [[g["value"], str(g["count"])] for g in groups]
        # Timeline?
        timeline = tool_result.get("timeline", [])
        if timeline:
            return ["Mês", "Contagem"], [[t[0], str(t[1])] for t in timeline]
        return [], []
    
    # Determinar headers a partir das keys do primeiro item
    sample = items[0]
    # Ordem preferencial
    preferred = ["id", "type", "title", "state", "area", "assigned_to", "created_by", "created_date", "url"]
    headers = [k for k in preferred if k in sample]
    headers.extend(k for k in sample.keys() if k not in headers and k != "score")
    
    rows = []
    for item in items:
        rows.append([str(item.get(h, "")) for h in headers])
    
    return headers, rows


def _clean_header(h: str) -> str:
    """Limpa header para display."""
    return h.replace("_", " ").title()


def _safe_sheet_title(title: str) -> str:
    """Sanitize Excel worksheet title (max 31 chars, no []:*?/\\)."""
    if not title:
        return "Export"
    forbidden = set('[]:*?/\\')
    safe = ''.join('_' if ch in forbidden else ch for ch in str(title))
    safe = safe.strip().strip("'")
    if not safe:
        safe = "Export"
    return safe[:31]


def _latin1_safe(text: str, max_len: int = 0) -> str:
    """Sanitize text for fpdf2 core fonts (Latin-1 only)."""
    if not text:
        return ""
    safe = text.encode('latin-1', 'replace').decode('latin-1')
    if max_len > 0:
        safe = safe[:max_len]
    return safe


# =============================================================================
# CSV EXPORT
# =============================================================================

def to_csv(tool_result: dict, filename: str = "export.csv") -> io.BytesIO:
    """Gera CSV (UTF-8 BOM para compatibilidade Excel)."""
    headers, rows = extract_table_data(tool_result)
    
    buf = io.BytesIO()
    buf.write(b'\xef\xbb\xbf')  # UTF-8 BOM
    
    wrapper = io.TextIOWrapper(buf, encoding='utf-8', newline='')
    writer = csv.writer(wrapper)
    writer.writerow([_clean_header(h) for h in headers])
    writer.writerows(rows)
    wrapper.flush()
    wrapper.detach()
    
    buf.seek(0)
    return buf


# =============================================================================
# XLSX EXPORT (openpyxl)
# =============================================================================

def to_xlsx(tool_result: dict, title: str = "Export", filename: str = "export.xlsx") -> io.BytesIO:
    """Gera XLSX formatado com branding Millennium."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return to_csv(tool_result, filename.replace(".xlsx", ".csv"))
    
    headers, rows = extract_table_data(tool_result)
    if not headers:
        return to_csv(tool_result)
    
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(title)
    
    # Branding colors
    brand_fill = PatternFill(start_color="CC0033", end_color="CC0033", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_font = Font(size=10)
    zebra_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{EXPORT_AGENT_NAME} — {title}")
    title_cell.font = Font(bold=True, size=14, color="CC0033")
    
    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub = ws.cell(row=2, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {tool_result.get('total_count', len(rows))} registos")
    sub.font = Font(size=9, color="666666", italic=True)
    
    # Headers (row 4)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=_clean_header(h))
        cell.font = header_font
        cell.fill = brand_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    for r_idx, row in enumerate(rows, 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = border
            if (r_idx - 5) % 2 == 1:
                cell.fill = zebra_fill
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(_clean_header(headers[col-1]))
        for row in rows[:50]:  # Sample for performance
            if col-1 < len(row):
                max_len = max(max_len, len(str(row[col-1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =============================================================================
# PDF EXPORT (fpdf2)
# =============================================================================

def to_pdf(tool_result: dict, title: str = "Export", summary: str = "") -> io.BytesIO:
    """Gera PDF com tabela de dados."""
    try:
        from fpdf import FPDF
    except ImportError:
        # Fallback
        buf = io.BytesIO()
        buf.write(b"PDF generation requires fpdf2. Install: pip install fpdf2")
        buf.seek(0)
        return buf

    try:
        headers, rows = extract_table_data(tool_result)

        pdf = FPDF()
        pdf.add_page('L')  # Landscape for tables
        pdf.set_auto_page_break(auto=True, margin=15)

        # Title
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_text_color(204, 0, 51)  # Brand red
        pdf.cell(0, 10, _latin1_safe(title, 80), ln=True)

        # Subtitle
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, _latin1_safe(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | {EXPORT_AGENT_NAME}"), ln=True)
        if summary:
            pdf.cell(0, 6, _latin1_safe(summary, 120), ln=True)
        pdf.ln(5)

        if not headers:
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 10, "Sem dados para exportar.", ln=True)
        else:
            # Calculate column widths
            page_w = pdf.w - 20  # margins
            col_w = min(page_w / len(headers), 50)
            col_widths = [col_w] * len(headers)
            # Make title column wider
            if "title" in headers:
                ti = headers.index("title")
                extra = page_w - col_w * len(headers)
                if extra > 0:
                    col_widths[ti] += extra

            # Header row
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(204, 0, 51)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 7, _latin1_safe(_clean_header(h), 20), border=1, fill=True)
            pdf.ln()

            # Data rows
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(0, 0, 0)
            for r_idx, row in enumerate(rows[:200]):  # Max 200 rows in PDF
                if r_idx % 2 == 1:
                    pdf.set_fill_color(245, 245, 245)
                    fill = True
                else:
                    fill = False
                for i, val in enumerate(row):
                    w = col_widths[i] if i < len(col_widths) else col_w
                    pdf.cell(w, 6, _latin1_safe(str(val), 60), border=1, fill=fill)
                pdf.ln()

        # Footer
        pdf.ln(10)
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 5, _latin1_safe(f"{EXPORT_BRAND_NAME} | {EXPORT_AGENT_NAME} v7.0"), ln=True)

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf
    except Exception as e:
        logging.error("[ExportEngine] to_pdf failed: %s", e)
        pdf_fallback = FPDF()
        pdf_fallback.add_page()
        pdf_fallback.set_font('Helvetica', '', 12)
        pdf_fallback.cell(0, 10, "Erro ao gerar PDF. Tenta CSV ou XLSX.", ln=True)
        buf = io.BytesIO()
        pdf_fallback.output(buf)
        buf.seek(0)
        return buf


# =============================================================================
# SVG CHART EXPORT
# =============================================================================

def to_svg_bar_chart(tool_result: dict, title: str = "Chart") -> str:
    """Gera SVG bar chart simples a partir de groups ou distribution."""
    groups = tool_result.get("groups", [])
    if not groups:
        # Try state_distribution
        dist = tool_result.get("state_distribution", tool_result.get("type_distribution", {}))
        if dist:
            groups = [{"value": k, "count": v} for k, v in sorted(dist.items(), key=lambda x: x[1], reverse=True)]
    
    if not groups:
        return '<svg xmlns="http://www.w3.org/2000/svg" width="400" height="100"><text x="20" y="50">Sem dados</text></svg>'
    
    groups = groups[:20]  # Max 20 bars
    max_val = max(g["count"] for g in groups) or 1
    
    bar_h = 28
    label_w = 180
    chart_w = 600
    bar_w = chart_w - label_w - 60
    total_h = len(groups) * bar_h + 80
    
    svg = [f'<svg xmlns="http://www.w3.org/2000/svg" width="{chart_w}" height="{total_h}" style="font-family:Arial,sans-serif">']
    
    # Title
    svg.append(f'<text x="{chart_w//2}" y="25" text-anchor="middle" font-size="16" font-weight="bold" fill="#CC0033">{title}</text>')
    
    y = 50
    for g in groups:
        w = int((g["count"] / max_val) * bar_w)
        label = str(g["value"])[:25]
        
        svg.append(f'<text x="{label_w - 10}" y="{y + 16}" text-anchor="end" font-size="11" fill="#333">{label}</text>')
        svg.append(f'<rect x="{label_w}" y="{y + 2}" width="{max(w, 2)}" height="{bar_h - 6}" fill="#CC0033" rx="3"/>')
        svg.append(f'<text x="{label_w + w + 8}" y="{y + 16}" font-size="11" fill="#666">{g["count"]}</text>')
        
        y += bar_h
    
    # Footer
    svg.append(f'<text x="{chart_w//2}" y="{total_h - 10}" text-anchor="middle" font-size="9" fill="#999">{EXPORT_AGENT_NAME} | {datetime.now().strftime("%d/%m/%Y")}</text>')
    svg.append('</svg>')
    
    return '\n'.join(svg)


# =============================================================================
# HTML REPORT EXPORT
# =============================================================================

def to_html_report(tool_result: dict, title: str = "Relatório", summary: str = "") -> str:
    """Gera relatório HTML completo com tabela e estilos."""
    headers, rows = extract_table_data(tool_result)
    
    html = f"""<!DOCTYPE html>
<html lang="pt">
<head><meta charset="UTF-8"><title>{title}</title>
<style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; }}
h1 {{ color: #CC0033; border-bottom: 3px solid #CC0033; padding-bottom: 10px; }}
.meta {{ color: #666; font-size: 0.9em; margin-bottom: 20px; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
th {{ background: #CC0033; color: white; padding: 10px 12px; text-align: left; font-size: 0.9em; }}
td {{ padding: 8px 12px; border-bottom: 1px solid #eee; font-size: 0.85em; }}
tr:nth-child(even) {{ background: #f9f9f9; }}
tr:hover {{ background: #fff0f0; }}
.footer {{ margin-top: 30px; font-size: 0.8em; color: #999; border-top: 1px solid #eee; padding-top: 10px; }}
a {{ color: #CC0033; text-decoration: none; }}
</style></head>
<body>
<h1>{title}</h1>
<div class="meta">Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {tool_result.get('total_count', len(rows))} registos{f' | {summary}' if summary else ''}</div>
"""
    if headers:
        html += '<table><thead><tr>'
        for h in headers:
            html += f'<th>{_clean_header(h)}</th>'
        html += '</tr></thead><tbody>'
        for row in rows:
            html += '<tr>'
            for i, val in enumerate(row):
                if headers[i] == 'url' and val.startswith('http'):
                    html += f'<td><a href="{val}" target="_blank">🔗 Link</a></td>'
                elif headers[i] == 'id' and len(row) > headers.index('url') if 'url' in headers else False:
                    url_val = row[headers.index('url')] if 'url' in headers else ''
                    html += f'<td><a href="{url_val}" target="_blank">{val}</a></td>'
                else:
                    html += f'<td>{val}</td>'
            html += '</tr>'
        html += '</tbody></table>'
    else:
        html += '<p>Sem dados tabulares para apresentar.</p>'
        # Show raw data
        html += f'<pre>{json.dumps(tool_result, indent=2, ensure_ascii=False)[:5000]}</pre>'
    
    html += f'\n<div class="footer">{EXPORT_BRAND_NAME} | {EXPORT_AGENT_NAME} v7.0</div>\n</body></html>'
    return html
